#openpyxl is the package thats required in python to deal with excel files
#pip install openpyxl

import openpyxl

book = openpyxl.load_workbook("/home/rits/Documents/SeleniumExcelPractise/PythonExcel.xlsx")
sheet = book.active # Get the active sheet 
# You can also access a specific sheet by name
# For example, if you want to access a sheet named 'Sheet1', you can use
#sheet = book['Sheet1']

cell = sheet.cell(row=1, column=1) # Accessing a specific cell
print(cell.value) # Print the value of the cell

#to write a value to a cell
sheet.cell(row=2, column=2).value = "Ashok" # Write data to a specific cell
book.save("/home/rits/Documents/SeleniumExcelPractise/PythonExcel.xlsx") # Save the changes to the workbook
cell = sheet.cell(row=2, column=2) # Access the cell again to see the updated value
print(cell.value)
#No of rows and columns in the sheet
print(sheet.max_row) # Get the maximum number of rows
print(sheet.max_column) # Get the maximum number of columns
# Iterate through all rows and columns

sheet['B3'].value = "Ram" # when referring B3 dont need to highlight the cell pointing to a sheet is enough.
book.save("/home/rits/Documents/SeleniumExcelPractise/PythonExcel.xlsx") # Save the changes to the workbook
print(sheet['B3'].value)

#print all the values in the sheet

for i in range(1, sheet.max_row + 1):                    #To get Rows
    for j in range(1, sheet.max_column + 1):            #To get Columns
        print(sheet.cell(row=i, column=j).value, end=' ')
    print()  # Print a new line after each row

print("--------------------------------------------------")

# To read all the values in a specific row, for example, row 3
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Test2":
        for j in range(2, sheet.max_column +1):      # If we need to get the data only and not test name start from 2.
            print(sheet.cell(row=i, column=j).value, end=' ')
    print()  

# In Test we handle the Excel in Dictonary format
#We store the data in a dictionary format, key is the first column and value is the rest of the columns
print("--------------------------------------------------")
Excel_Dict = {}
temp = {}

for i in range(1, sheet.max_row +1):
    for j in range(1, sheet.max_column + 1):
        Excel_Dict.update[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        #Excel_dict[here we gave the Key details] = here we gave the value details

print(Excel_Dict)

#Now i need to print the dictonary in which i need data from col 2 not the testcase name.
print("--------------------------------------------------")

Excel_Dict1 = {}
for i in range(1, sheet.max_row + 1):
    for j in range(2, sheet.max_column + 1):  # Start from column 2 to skip the first column
        Excel_Dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Excel_Dict1)