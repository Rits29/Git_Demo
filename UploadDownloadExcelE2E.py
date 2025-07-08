import openpyxl
from selenium.webdriver.common.by import By
from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions

chromedriver_autoinstaller.install()
driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
print("Website opened successfully")

#Dynamic Xpath
fruitName = "Apple"
price_col = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
print(price_col)
old_cost = driver.find_element(By.XPATH,"//div[text()='"+fruitName+"']/parent::div/parent::div/div["+price_col+"]/div").text
print(old_cost)

driver.find_element(By.ID, "downloadButton").click()  # Click on the download button to download file.
print("Download button clicked")
# Wait for the download to complete (you may need to adjust the wait time based on your system)
#Selenium doesnot handle windows only webelements are there. 
# So Selenium has given type attribute in the download button to handle the download. 
# If not given then we need to use AutoIT or Robot class to handle the download.

#We will edit the value of price via automation and upload the changed excel file.
book = openpyxl.load_workbook("/home/rits/Downloads/download.xlsx")
sheet = book.active

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1): 
        if sheet.cell(row=i, column=j).value == "Apple":
            row_index = i
            break

for j in range(1, sheet.max_column + 1):
    if sheet.cell(row =1, column = j).value == "price":
        col_index = j
        break

sheet.cell(row= row_index, column = col_index).value = "848"
book.save("/home/rits/Downloads/download.xlsx")

#upload the file. here we send the file path to the choose file by using sendkeys method.
file_upload = driver.find_element(By.ID, "fileinput")
file_upload.send_keys("/home/rits/Downloads/download.xlsx")  # Provide the path to the file you want to upload
print("File uploaded successfully")

wait = WebDriverWait(driver, 20)
flashmsg = (By.XPATH, "//div[@class='Toastify__toast-body']/div[2]")
wait.until(expected_conditions.visibility_of_element_located(flashmsg))
print(driver.find_element(*flashmsg).text)

#Dynamic Xpath 

actual_cost = driver.find_element(By.XPATH,"//div[text()='"+fruitName+"']/parent::div/parent::div/div["+price_col+"]/div").text
print(actual_cost)

driver.quit()
